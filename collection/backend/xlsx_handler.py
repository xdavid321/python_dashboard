import openpyxl
from openpyxl.utils import range_boundaries
import copy

class XLSX:

    def readSheet(self, file_location):
        wb = openpyxl.load_workbook(filename = file_location)

        ws = wb.worksheets[0]
        merged_cells = ws.merged_cells
        merged_cell_ranges = copy.deepcopy(merged_cells.ranges)
        for ranges in merged_cell_ranges:
            m_col, m_row, x_col, x_row = range_boundaries(str(ranges))
            cell_value = ""
            for row in ws.iter_rows(min_col = m_col, min_row=  m_row, max_col = x_col, max_row = x_row):
                for cell in row:
                    if cell.value != None:
                        cell_value = cell.value
            ws.unmerge_cells(str(ranges))
            for row in ws.iter_rows(min_col = m_col, min_row=  m_row, max_col = x_col, max_row = x_row):
                for cell in row:
                    cell.value = cell_value

        max_col = ws.max_column
        max_row = ws.max_row

        data = []
        for row in ws.iter_rows(min_col = 1, min_row = 1, max_col = ws.max_column, max_row = ws.max_row):
            data_col = []
            for cell in row:
                data_col.append(cell.value)
            data.append(data_col)
        return data

    def console_print_data(self, data):
        for row in data:
            for cell in row:
                print(cell, end=" ")
            print("")

    def process_data(self, data, header_end):
        headers = []
        for i in range(0, header_end):
            headers.append(data[0])
            data.pop(0)
        final_header_row = []
        for i, header_row in enumerate(headers):
            for j, head in enumerate(header_row):
                if(i==0):
                    final_header_row.append(str(head).lower().replace(" ", "_"))
                else:
                    if(final_header_row[j] != head):
                        if(final_header_row[j] != 'market__market'):
                            final_header_row[j] += "__" + str(head.lower().replace(" ", "_"))
        return headers, final_header_row, data

    def insert_into_db(self, raw_data, header_rows):
        headers, data = self.process_data(data=raw_data, header_end=header_rows)